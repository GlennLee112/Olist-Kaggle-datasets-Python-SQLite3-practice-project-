from openpyxl import *
from openpyxl import utils


def Excel_Write(xlsx, df, sheets_no, row_start=1, column_start=1):
    wb = load_workbook(xlsx)
    ws = wb.worksheets[sheets_no]

    for idx, row in df.iterrows():
        for col in range(len(row)):
            column_letter = utils.get_column_letter(column_start + col)
            ws[f'{column_letter}{idx + row_start + 1}'] = df.iloc[idx, col]

    wb.save(xlsx)


# def opx_writer_write(xlsx, df, )
